from openpyxl import Workbook
from openpyxl import load_workbook




wb = load_workbook(filename='done_estimate_analysis.xlsx',data_only=True)
main = wb['done_estimate_analysis']
infill = []
ep = []
numdecks= []
decklen=[]
line_post = []
out_corner =[]
in_corner = []
totals = []
accessor = []
estimate = []
categories= [infill,ep,numdecks,decklen,line_post,out_corner,in_corner,totals,accessor,estimate]

new_infill = []
new_ep = []
new_numdecks = []
new_decklen = []
new_line_post = []
new_out = []
new_in = []
new_total = []
new_accessor = []
new_categories = []
new_estimate=[]
for num in range(1,11):
    for col in main.iter_cols(min_col=num,max_col=num,min_row=1,max_row=997,values_only=True):
        for item in col:
            categories[num-1].append(item)


for num in range(len(infill)):
    if new_infill != 'None':
        if numdecks[num] != 'None':
            if decklen[num] != 'None':
                new_infill.append(infill[num])
                new_ep.append(ep[num])
                new_numdecks.append(numdecks[num])
                new_decklen.append(decklen[num])
                new_line_post.append(line_post[num])
                new_out.append(out_corner[num])
                new_in.append(in_corner[num])
                new_total.append(totals[num])
                new_accessor.append(accessor[num])
                new_estimate.append(estimate[num])
                
f= open('cleaned_data.csv', 'w')
for num in range(len(new_infill)):
    f.write(str(new_infill[num]) + ',' + str(new_ep[num]) + ',' + str(new_numdecks[num]) + ',' + str(new_decklen[num]) + ',' + str(new_line_post[num]) + ',' + str(new_out[num]) + ',' + str(new_in[num]) + ',' + str(new_total[num])+ ',' + str(new_accessor[num]) + ',' + str(new_estimate[num]) + '\n')
                    
f.close()
