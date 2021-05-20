import os
from openpyxl import Workbook
from openpyxl import load_workbook

os.chdir('/Users/Owner/Desktop/done')
estimate_dict = {}

estimates = []
old = []
failed = []
ep = []
line_posts = []
out_corner = []
in_corner = []

deck_len = []
deck_q = []
totals = []
items_array = [line_posts,ep,in_corner,out_corner]
index = []

print(os.getcwd())
rail_type = []

def search_row(_area,_col,_minrow,_maxrow,_array,num=0):
    for row in _area.iter_rows(min_col=_col + num,max_col=_col + num,min_row=_minrow,max_row=_maxrow,values_only=True):
        res = ''.join(map(str,row))
        _array.append(str(res))


#print(os.listdir())
#print('------------------------------------')
for files in os.listdir():
    print(files)
    filez = os.chdir('/Users/Owner/Desktop/done/' + str(files))
    #print(filez)
    for file in os.listdir():
        #print(file)
        if file.endswith('xlsm'):
            print('found estimate...')
            try:
                wb = load_workbook(filename=file,data_only=True)
                main_a = wb['Main']
                area_a = wb['Area A']
                area_b = wb['Area B']
                area_c = wb['Area C']
                area_d = wb['Area D']
                areas = [area_a,area_b,area_c,area_d]
 
                
                for row in main_a.iter_rows(min_col=14,max_col=14,min_row=28,max_row=31,values_only=True):
                    res = ''.join(map(str,row))
                    deck_len.append(str(res))
                    
                for row in main_a.iter_rows(min_row=28,max_row=31,min_col=15,max_col=15,values_only=True):
                    res = ''.join(map(str,row))
                    deck_q.append(str(res))
                    
                for num in range(4):
                    search_row(main_a,9,28,31,items_array[num],num)

                    index.append((len(estimates) + 1))

                    for row in main_a.iter_rows(min_col=20,max_col=20,min_row=36,max_row=36,values_only = True):
                        res=''.join(map(str,row))
                        totals.append(str(res))

                    
                for area in areas:
                    add = 0
                    for row in area.iter_rows(min_row=28,max_row=36,min_col=2,max_col=2,values_only=True):
                        res = ''.join(map(str,row))
                        print(res)
                        if 'CTG' in str(res):
                            rail_type.append('glass')
                            add +=1
                            break
                        elif 'CTLG' in str(res):
                            rail_type.append('glass')
                            add +=1
                            break
                        elif 'PT1-420' in str(res):
                            rail_type.append('picket')
                            add +=1 
                            break
                        elif 'SS Cable' in str(res):
                            rail_type.append('cable')
                            add +=1
                            break
                        else:
                            pass
                    if add == 0:
                        rail_type.append('NA')
                                     
                old.append(file)
                estimates.append(str(files) + "/" + str(file))
                print('SUCCESS - OLD')
            except:
                print('not old format...')
                print('try new format...')
                
            try:
                wb = load_workbook(filename=file,data_only=True)
                est_total = wb['Estimate Total']
                pro_materials = wb['Project Materials']
                area_a = wb['Area A']
                area_b = wb['Area B']
                area_c = wb['Area C']
                area_d = wb['Area D']
                areas = [area_a,area_b,area_c,area_d]
                for row in pro_materials.iter_rows(min_row=25,max_row=28,min_col=15,max_col=15,values_only=True):
                    res = ''.join(map(str,row))
                    deck_len.append(str(res))
                for row in pro_materials.iter_rows(min_row=25,max_row=28,min_col=16,max_col=16,values_only=True):
                    res = ''.join(map(str,row))
                    deck_q.append(str(res))
                    
                for num in range(4):
                    search_row(pro_materials,10,25,28,items_array[num],num)


                    index.append((len(estimates) + 1))
                    
                    for row in est_total.iter_rows(min_row=5,max_row=5,min_col=20,max_col=20,values_only=True):
                        res = ''.join(map(str,row))
                        totals.append(str(res))
                 
                for area in areas:
                    add = 0
                    for row in area.iter_rows(min_row=28,max_row=36,min_col=2,max_col=2,values_only=True):
                        res = ''.join(map(str,row))
                        print(res)
                        if 'CTG' in str(res):
                            rail_type.append('glass')
                            add += 1
                            break
                        elif 'CTLG' in str(res):
                            rail_type.append('glass')
                            add +=1
                            break
                        elif 'PT1-420' in str(res):
                            rail_type.append('picket')
                            add += 1
                            break
                        elif 'SS Cable' in str(res):
                            rail_type.append('cable')
                            add +=1
                            break
                        else:
                            pass
                    if add == 0:
                        rail_type.append('NA')
                estimates.append(str(files) + "/" + str(file))
                print('SUCCESS - NEW')
                    
            except:
                print('not proper format...')
                failed.append(str(files) + "/" + str(file))
                    
        
#print(estimates)
print(ep)
print(rail_type)
#print(old)

print()
print('total of {} old estimates found and {} total estimates found!'.format(len(old),len(estimates)))
print('{} failed to load...'.format(len(failed)))
print()
print('# of section of railing types: ' + str(len(rail_type)))
print('# of sections of end posts: ' + str(len(ep)))
print('# of sections of deck quantuies: ' + str(len(deck_q)))
print('# of deck length sections: ' + str(len(deck_len)))
print('# of outside corner posts : ' + str(len(out_corner)))
print('# of inside corner posts: ' + str(len(in_corner)))
print('# of line posts: ' + str(len(line_posts)))
print('# of total prices: ' + str(len(totals)))
file = open('done_scrape.csv','w')
file.write('Infill, ep,deck_q,deck_len,line_post,our corner,in_corner,totals,index,estimate name \n')
for item in range(len(rail_type)):
    file.write( str(rail_type[item]) + ',' + str(ep[item])+ ',' + str(deck_q[item]) + ',' + str(deck_len[item]) + ',' + str(line_posts[item]) + ',' + str(out_corner[item]) + ',' + str(in_corner[item]) + ',' + str(totals[item]) +','+ str(index[item]) + ',' + str(estimates[int(index[item] -1)]) + '\n')

file.close()
